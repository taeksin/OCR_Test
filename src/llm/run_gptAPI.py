"""OpenAI Chat Completions API 호출 예제 (한글 주석/출력)

이 모듈은 .env에서 API 키를 읽어 OpenAI 클라이언트를 생성한 뒤,
사용자 프롬프트를 GPT 모델에 전달하고 응답 텍스트를 출력합니다.
"""

import os

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from tqdm import tqdm  # tqdm 모듈 추가
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed


def load_api_key(env_key_name="OPENAI_API_KEY"):
    """.env 또는 환경 변수에서 OpenAI API 키를 불러옵니다.

    Args:
        env_key_name: 읽어올 환경 변수 이름 (기본: OPENAI_API_KEY)

    Returns:
        API 키 문자열

    Raises:
        RuntimeError: API 키를 찾지 못한 경우
    """
    load_dotenv()
    api_key = os.getenv(env_key_name)
    if not api_key:
        raise RuntimeError(f"환경 변수 '{env_key_name}'를 찾지 못했습니다. .env에 설정해 주세요.")
    return api_key


def create_client(api_key):
    """제공된 API 키로 OpenAI 클라이언트를 생성합니다."""
    return OpenAI(api_key=api_key)


def run_gpt5(prompt):
    """주어진 프롬프트를 GPT 모델에 전달하고 응답 텍스트를 반환합니다."""
    api_key = load_api_key()
    client = create_client(api_key)

    # 사용 가능한 모델명은 계정/권한에 따라 다를 수 있습니다.
    model_name = "gpt-5"

    completion = client.chat.completions.create(
        model=model_name,
        messages=[
            {"role": "system", "content": "당신은 도움을 주는 어시스턴트입니다."},
            {"role": "user", "content": prompt},
        ],
        temperature=1,
        max_completion_tokens=2024,
        reasoning_effort="minimal",
    )

    return completion.choices[0].message.content or ""


def extract_first_json(text):
    """응답 문자열에서 첫 번째 유효한 JSON 객체만 추출하여 문자열로 반환.

    - ```json ... ``` 코드블록이 있으면 내부만 파싱 시도
    - 텍스트에 섞여 있을 경우 첫 번째 { ... } 범위를 찾아 파싱
    """
    if not text:
        return ""

    code_block = re.search(r"```(?:json)?\s*([\s\S]*?)```", text, re.IGNORECASE)
    if code_block:
        candidate = code_block.group(1).strip()
        try:
            parsed = json.loads(candidate)
            return json.dumps(parsed, ensure_ascii=False)
        except Exception:
            pass

    brace_stack = []
    start_idx = -1
    for idx, ch in enumerate(text):
        if ch == '{':
            if not brace_stack:
                start_idx = idx
            brace_stack.append('{')
        elif ch == '}':
            if brace_stack:
                brace_stack.pop()
                if not brace_stack and start_idx != -1:
                    candidate = text[start_idx:idx + 1]
                    try:
                        parsed = json.loads(candidate)
                        return json.dumps(parsed, ensure_ascii=False)
                    except Exception:
                        start_idx = -1
                        continue

    return ""


def process_excel(file_path):
    """B열과 D열의 텍스트를 LLM에 보내고, 결과를 각각 C열과 E열에 저장합니다.

    - 헤더는 1행에 있다고 가정하고 2행부터 처리합니다.
    - Bn -> Cn, Dn -> En 으로 기록합니다.
    """
    wb = load_workbook(filename=file_path)
    ws = wb.active

    instruction = "키-값 쌍으로 JSON 뽑아줘"

    # 2행부터 마지막 행까지 반복: 요청 목록 준비
    tasks = []  # (row_idx, target_col, prompt)
    for row_idx in tqdm(range(2, ws.max_row + 1), desc="Preparing rows", unit="row"):
        b_val = ws.cell(row=row_idx, column=2).value  # B열
        d_val = ws.cell(row=row_idx, column=4).value  # D열

        b_text = str(b_val).strip() if b_val is not None else ""
        d_text = str(d_val).strip() if d_val is not None else ""

        if b_text:
            b_prompt = f"{instruction}\n\n원문:\n{b_text}"
            tasks.append((row_idx, 3, b_prompt))  # C열에 기록

        if d_text:
            d_prompt = f"{instruction}\n\n원문:\n{d_text}"
            tasks.append((row_idx, 5, d_prompt))  # E열에 기록

    # 병렬 LLM 호출
    if tasks:
        max_workers = max(4, (os.cpu_count() or 4))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_cell = {
                executor.submit(run_gpt5, prompt): (row_idx, col)
                for (row_idx, col, prompt) in tasks
            }

            for future in tqdm(as_completed(future_to_cell), total=len(future_to_cell), desc="LLM calls", unit="req"):
                row_idx, col = future_to_cell[future]
                try:
                    resp = future.result()
                    ws.cell(row=row_idx, column=col).value = extract_first_json(resp)
                except Exception as e:
                    ws.cell(row=row_idx, column=col).value = f"ERROR: {e}"

    wb.save(filename=file_path)


def main():
    # 엑셀 파일 경로 (현재 파일 기준 동일 폴더)
    excel_path = os.path.join(os.path.dirname(__file__), "LLM API결과 테스트.xlsx")
    try:
        process_excel(excel_path)
        print(f"엑셀 처리 완료: {excel_path}")
    except Exception as e:
        print(f"오류: {e}")


if __name__ == "__main__":
    main()
